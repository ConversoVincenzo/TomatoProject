from openpyxl import Workbook, load_workbook
import calendar
import time
from pydub import AudioSegment
from pydub.playback import play

path = "work_load.xlsx"

workload = load_workbook(path)
sheet = workload.active
slotsNum = sheet.max_column
row = sheet.max_row
toDoList = []

#To do list setup
for col in range(slotsNum):
    quarter = []
    for rw in range(row):
        temp = sheet.cell(row=rw+1, column=col+1).value
        if temp is None:
            break
        quarter.append(temp)
    toDoList.append(quarter)

minutes = int(input("Choose quarter lenght [in minutes]: "))

print(toDoList)

print("Wait ...")
result = Workbook()
sheet = result.active

sheet["A1"] = "Slots"
sheet["B1"] = "In Time?"

sound = AudioSegment.from_wav('.\\audiofile.wav')

input("ATTENTION! The sound for the timer may be to loud, control your mixer than press [Enter] to listen the sound. ")

play(sound)

for slot in range(len(toDoList)):
    print("For this slot you need to " + str(toDoList[slot]) + "\n")
    input("The Timer will start when you press [Enter]")
    print("GL HF!")
    time.sleep(minutes * 60)
    print("Ding!!!")
    play(sound)
    stat = input("Did you do it on time? Y or N: ").upper()
    while stat != "Y" and stat != "N":
        stat = input("Bro be serious please. Y or N: ")
    sheet.cell(row=slot + 1, column=1).value = str(toDoList[slot])
    sheet.cell(row=slot + 1, column=2).value = stat
    if slot%3 == 1:
        print("time to touch some grass, come back after 15 minutes")
        time.sleep(15 * 60)
    else:
        print("time to touch some grass, come back after 5 minutes")
        time.sleep(5 * 60)


print("Ending session ...")

current_GMT = time.gmtime()

time_stamp = calendar.timegm(current_GMT)

result.save("files/result" + str(time_stamp) + ".xlsx")